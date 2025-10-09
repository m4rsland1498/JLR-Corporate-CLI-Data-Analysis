import requests
from bs4 import BeautifulSoup
import os
import openpyxl as pyxl
from simple_term_menu import TerminalMenu
from termgraph import Data, Args, BarChart, VerticalChart, Colors
import termcharts

##########################################################################################
        
def downloadData(year, quarter):
    year = str(year)[2:]
    
    url = "https://www.jlr.com/results-centre"
    try:
        response = requests.get(url)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Failed to retrieve the webpage: {e}")
        return
    soup = BeautifulSoup(response.text, "html.parser")
    
    year_tab = soup.find("li", string=lambda text: text and text.strip() == f"FY{year}")
    if year_tab is None:
        print(f"Could not find the tab for FY{year}.")
        return
    year_id = year_tab.get("data-year-id")
    results_table = soup.find("section", attrs={"data-table-id": year_id})
    
    link_element = results_table.find(attrs={"aria-label": f"Download Sales Volumes Q{quarter} file"})
    
    if link_element:
        file_url = link_element.get("href")
        print(f"Found download link: {file_url}")
    else:
        print("Could not find sales data.")
        file_url = None
        
    if file_url:
        try:
            file_name = os.path.basename(file_url.split("?")[0])
            download_folder = "sales_data_cache"
            if not os.path.exists(download_folder):
                os.makedirs(download_folder)
            
            save_path = os.path.join(download_folder, file_name)
            
            print(f"Downloading {file_name}...")
            response = requests.get(file_url, stream=True) # stream=True just handles more efficiently
            response.raise_for_status() # successful status code is 200
            
            with open(save_path, "wb") as file:
                for chunk in response.iter_content(chunk_size=8192):
                    file.write(chunk)
                    
            print(f"Download complete. File cached to: {save_path}")
            
        except:
            print(f"Error occured: {requests.exceptions.RequestException}")
            
        return save_path

##########################################################################################
            
def parseMenu(save_path):
    df = pyxl.load_workbook(save_path, data_only=True)
    global df1
    try:
        df1 = df["JLR Retails to Date"]
    except:
        df1 = df["Website Retails"]
    
    col_a = next(df1.iter_cols(min_col=1, max_col=1))
    col_b = next(df1.iter_cols(min_col=2, max_col=2))
    active_col = col_a
    
    global brands_and_models
    brands_and_models = []
    temp = []
    global attempts
    attempts = 0
    firstCol = False
    
    while brands_and_models == []:
        attempts+=1
        for row in range(0,df1.max_row):
            data = active_col[row].value
            if ("Retail" in str(data)
                or "CJLR" in str(data)
                or "No longer" in str(data)
                or "Note" in str(data)
                or "0" in str(data)
                or "Brand" in str(data)
                or "Alternative" in str(data)):
                data = None
            brands_and_models.append(data)
        
        for i in brands_and_models:
            if i != None and i!="*":
                firstCol = True
                
        if not firstCol:
            brands_and_models = []
        active_col = col_b
        
    temp = []
    for i in brands_and_models:
        if i != None:
            temp.append(str(i))
        else:
            temp.append(i)
    brands_and_models = temp
        
##########################################################################################
        
def parseAndPresentData(selected_index, save_path):
    global df1
    global year
    global attempts
    
    full_labels = [f"QTD{year}", f"QTD{year-1}",
                   f"FYTD{year}", f"FYTD{year-1}",
                   f"CYTD{year}", f"CYTD{year-1}"]
    colours = [Colors.Magenta, Colors.Green, Colors.Blue, Colors.Red, Colors.Yellow, Colors.Cyan, None]

    for i in range(len(selected_index)):
        row_values = []
        non_empty_count = 0
        full_row = df1[selected_index[i] + 1]
        
        for cell in full_row:
            cell_value = cell.value
            if cell_value is not None and cell_value != '':
                is_numeric = isinstance(cell_value, (int, float))
                
                if is_numeric:
                    non_empty_count += 1
                    # append the value if it's a number and not the third one in a row
                    if non_empty_count % 3 != 0:
                        row_values.append([float(cell_value)])
                else:
                    # reset the count if a non-number in cell - accounts for errors in excel file as I found a div/0 error
                    non_empty_count = 0
        
        labels_to_use = full_labels[:len(row_values)]
        presentData = Data(row_values, labels_to_use)
        args = Args(
            title=f"JLR Retails to Date - {df1.cell(row=selected_index[i]+1, column=attempts).value}",
            width=80,
            no_readable=True,
            colors=[colours[i % 7]],
            format="{:.0f}",
        )
        
        chart = BarChart(presentData, args)
        chart.draw()
        
        print([item[0] for item in row_values], "\n\n")

        #pie charts
        # data needs to be grabbed differently to support older excel file formatting
        qpoySectionName = df1.cell(row=selected_index[i]+1,column=attempts).value
        qpoySectionValue = df1.cell(row=selected_index[i]+1,column=1+attempts).value
        qpofyTotalSubSection = df1.cell(row=selected_index[i]+1,column=5).value - qpoySectionValue
        qpocyTotalSubSection = df1.cell(row=selected_index[i]+1,column=8).value - qpoySectionValue
        qpofyData = {qpoySectionName:qpoySectionValue, "Remaining Fiscal Year Sales - "+qpoySectionName:qpofyTotalSubSection}
        qpocyData = {qpoySectionName:qpoySectionValue, "Remaining Fiscal Year Sales - "+qpoySectionName:qpocyTotalSubSection}
        if qpofyTotalSubSection > 0:
            qpofyChart = termcharts.pie(qpofyData,
            title = "Quarter's Percentage of FYTD") # quarter's percentage of fiscal year to date
            print(qpofyChart)
        else:
            print("This quarter's sales make up all of the fiscal year to date.")
        if qpocyTotalSubSection > 0:
            qpocyChart = termcharts.pie(qpocyData,
            title = "Quarter's Percentage of CYTD") # quarter's percentage of calendar year to date
            print(qpocyChart)
        else:
            print("This quarter's sales make up all of the calendar year to date.")
        

    os.remove(save_path)
        
##########################################################################################

def menu(save_path):
    global brands_and_models
    options = brands_and_models
    terminal_menu = TerminalMenu(options, skip_empty_entries=True,
                                 show_search_hint=True, multi_select=True,
                                 show_multi_select_hint=True)
    selected_index = terminal_menu.show()
    parseAndPresentData(selected_index, save_path)
    

##########################################################################################
    


##########################################################################################

os.system("clear")
global year
while True:
    FYs = {2022:[1,2,3,4], 2023:[1,2,3,4], 2024:[1,2,3,4], 2025:[1,2,3,4], 2026:[1]}
    
    print("\n\nSelect a Fiscal Year")
    
    years = list(FYs.keys())
    years_str = [str(y) for y in years]
    year_menu = TerminalMenu(years_str, show_search_hint=True)
    year_index = year_menu.show()
    year = years[year_index]
    os.system("clear")
    
    print("\nSelect a Quarter")
    
    quarters = FYs[year]
    quarter_menu = TerminalMenu([str(q) for q in quarters], show_search_hint=True)
    quarter_index = quarter_menu.show()
    quarter = quarters[quarter_index]
    os.system("clear")
                
    save_path = downloadData(year, quarter)
    parseMenu(save_path)
    menu(save_path)
